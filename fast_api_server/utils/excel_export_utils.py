import io

import openpyxl as xl
from openpyxl.utils import get_column_letter

from shared.models import ClassModel
from shared.db import database
from shared.books_repository import BooksRepository

db = database.get_database()
book_repo = BooksRepository()

async def write_results_to_xlsx(klass: ClassModel) -> io.BytesIO:
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    ws.title = klass.name

    students = klass.students

    # write class list (usernames)
    ws.cell(row=2, column=2).value = 'username'
    ws.column_dimensions[get_column_letter(2)].width = 10
    for i, student in enumerate(students):
        ws.cell(row=i+3, column=2).value = student

    ws.row_dimensions[2].height = 80

    # resolved names cache
    names = {}

    # total tests passed this year
    grand_totals = {}

    # set up border style, so we can reuse this for all result cells
    border = xl.styles.Border(left=xl.styles.Side(style='thin'), right=xl.styles.Side(style='thin'), top=xl.styles.Side(style='thin'), bottom=xl.styles.Side(style='thin'))

    # for each book and each student, populate the sheet with results
    col = 3
    for book_name in klass.books:
        book = book_repo.load_book(book_name)
        if not book:
            continue
        testable_nodes = book_repo.book_to_testable_node_list(book) 
        book_res = await db.get_results_for_users(book_name, klass.students)
        for i, node in enumerate(testable_nodes):
            cell = ws.cell(row=2, column=col+i)
            cell.value = node.name.replace('Challenge', 'C').strip()
            cell.alignment = xl.styles.Alignment(textRotation=90)
            ws.column_dimensions[get_column_letter(col+i)].width = 2.5
        cell = ws.cell(row=2, column=col+len(testable_nodes))
        cell.value = 'total in book'
        cell.alignment = xl.styles.Alignment(textRotation=90)
        ws.column_dimensions[get_column_letter(col+len(testable_nodes))].width = 4
       

        for i, student in enumerate(students):
            res = [x for x in book_res if x['user'] == student]
            if len(res) != 1:
                for iNode, node in enumerate(testable_nodes):
                    ws.cell(row=i+3, column=col+iNode).value = ' '
                    ws.cell(row=i+3, column=col+iNode).border = border
                continue
            res = res[0]
            if res.get('name') and student not in names:
                names[student] = res.get('name')
            passed = 0
            for iNode, node in enumerate(testable_nodes):
                if node.id not in res:
                    ws.cell(row=i+3, column=col+iNode).value = ' '
                    ws.cell(row=i+3, column=col+iNode).border = border
                    continue
                if isinstance(res[node.id], bool):
                    if res[node.id]:
                        passed += 1
                        ws.cell(row=i+3, column=col+iNode).value = 'y'
                        ws.cell(row=i+3, column=col+iNode).border = border
                    else:
                        ws.cell(row=i+3, column=col+iNode).value = 'n'
                        ws.cell(row=i+3, column=col+iNode).border = border
                else:
                    correct = res[node.id].get('correct', False)
                    wrong_attempts = res[node.id].get('wrong_attempts', 0)
                    if correct and wrong_attempts == 0:
                        passed += 1
                        ws.cell(row=i+3, column=col+iNode).value = 'y'
                        ws.cell(row=i+3, column=col+iNode).border = border
                    elif correct and wrong_attempts > 0:
                        passed += 1
                        ws.cell(row=i+3, column=col+iNode).value = 'f'
                        ws.cell(row=i+3, column=col+iNode).border = border
                    else:
                        ws.cell(row=i+3, column=col+iNode).value = 'n'
                        ws.cell(row=i+3, column=col+iNode).border = border
            ws.cell(row=i+3, column=col+len(testable_nodes)).value = passed
            ws.cell(row=i+3, column=col+len(testable_nodes)).border = border
            grand_totals[student] = grand_totals.get(student, 0) + passed              

        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+len(testable_nodes))
        ws.cell(row=1, column=col).alignment = xl.styles.Alignment(horizontal='center')
        ws.cell(row=1, column=col).value = book.name
        col += len(testable_nodes) + 1

    # grand totals
    cell = ws.cell(row=2, column=col)
    cell.value = 'grand total'
    cell.alignment = xl.styles.Alignment(textRotation=90)
    ws.column_dimensions[get_column_letter(col)].width = 4
    for i, student in enumerate(students):
        ws.cell(row=i+3, column=col).value = grand_totals.get(student, 0)
        ws.cell(row=i+3, column=col).border = border

    # color scale for grand totals
    ws.conditional_formatting.add(f'{get_column_letter(col)}3:{get_column_letter(col)}{len(students)+2}', 
                                  xl.formatting.rule.ColorScaleRule(start_type='percentile', start_value=0, start_color='FFF8696B',
                                    mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                                    end_type='percentile', end_value=90, end_color='FF63BE7B'))

    # add resolved names to class list
    ws.column_dimensions[get_column_letter(1)].width = 15
    ws.cell(row=2, column=1).value = 'names'
    for i, student in enumerate(students):
        ws.cell(row=i+3, column=1).value = names.get(student, student)

    
    # set up conditional formatting
    def make_rule(color: str, text: str):
        dxf = xl.styles.differential.DifferentialStyle(fill=xl.styles.PatternFill(bgColor=color), font=xl.styles.Font(color=color))
        rule = xl.formatting.Rule(type='containsText', dxf=dxf, text=text, operator='containsText')
        rule.formula = [f'NOT(ISERROR(SEARCH("{text}", C3)))']
        return rule
    ws.conditional_formatting.add('C3:ZZ1000', make_rule('aa0000', 'n'))
    ws.conditional_formatting.add('C3:ZZ1000', make_rule('00ee00', 'y'))
    ws.conditional_formatting.add('C3:ZZ1000', make_rule('00bb00', 'f'))
    
    # freeze panes
    ws.freeze_panes = ws['C2']

    # save
    buf = io.BytesIO()
    wb.save(buf)
    return buf
